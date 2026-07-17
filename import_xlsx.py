# -*- coding: utf-8 -*-
"""
import_xlsx.py

Importa fisiere Excel (.xlsx) cu format multi-sezon, de tipul celor descarcate
de pe football-data.co.uk in format agregat (o singura foaie cu toate sezoanele
unei tari/campionat, gen "Romania_totale.xlsx").

Coloane asteptate in foaia de calcul (header pe primul rand):
    Country, League, Season, Date, Time, Home, Away, HG, AG, Res,
    PSCH, PSCD, PSCA (Pinnacle)
    MaxCH, MaxCD, MaxCA (cota maxima de piata)
    AvgCH, AvgCD, AvgCA (cota medie de piata) <- FOLOSITA implicit pentru odds
    BFECH, BFECD, BFECA (Betfair Exchange)
    B365CH, B365CD, B365CA (Bet365)

Foloseste coloanele Avg* (cota medie de piata) ca sursa implicita pentru
odds_home / odds_draw / odds_away, considerata cea mai reprezentativa
(nedistorsionata de un singur bookmaker).

Cum se foloseste (din terminal, in folderul proiectului):

    python import_xlsx.py cale\\catre\\fisier.xlsx "Superliga" club_league Romania

Argumente, in ordine:
    1. calea catre fisierul .xlsx
    2. numele competitiei in baza de date (trebuie sa existe deja, vezi
       database.py -> seed_initial_competitions)
    3. tipul competitiei: club_league / national_team / european_cup
    4. tara (sau "None" daca nu se aplica)

Sezonul se ia automat din coloana "Season" a fisierului, deci un singur
fisier poate importa mai multe sezoane deodata.

Necesita biblioteca openpyxl:
    pip install openpyxl
"""

import sys
from datetime import datetime, date

import openpyxl

from database import get_connection, get_or_create_team, get_or_create_competition, create_tables

# Numele coloanelor din fisierul Excel (ajusteaza aici daca fisierul tau difera)
COL_COUNTRY = "Country"
COL_LEAGUE = "League"
COL_SEASON = "Season"
COL_DATE = "Date"
COL_HOME = "Home"
COL_AWAY = "Away"
COL_HG = "HG"
COL_AG = "AG"
COL_ODDS_HOME = "AvgCH"
COL_ODDS_DRAW = "AvgCD"
COL_ODDS_AWAY = "AvgCA"


def to_iso_date(value) -> str:
    """Converteste o valoare de data (datetime, date, sau text) la format ISO YYYY-MM-DD."""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, str) and value.strip():
        for fmt in ("%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d"):
            try:
                return datetime.strptime(value.strip(), fmt).strftime("%Y-%m-%d")
            except ValueError:
                continue
    return None


def safe_float(value):
    try:
        if value is None or value == "":
            return None
        return float(value)
    except (ValueError, TypeError):
        return None


def safe_int(value):
    try:
        if value is None or value == "":
            return None
        return int(value)
    except (ValueError, TypeError):
        return None


def import_xlsx(filepath: str, competition_name: str, competition_type: str, country: str):

    create_tables()

    competition_id = get_or_create_competition(competition_name, competition_type, country)

    team_type = "national" if competition_type == "national_team" else "club"

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active  # prima foaie activa; daca ai nevoie de alta, ajusteaza aici

    header = [c.value for c in ws[1]]
    try:
        idx = {name: header.index(name) for name in [
            COL_SEASON, COL_DATE, COL_HOME, COL_AWAY, COL_HG, COL_AG,
            COL_ODDS_HOME, COL_ODDS_DRAW, COL_ODDS_AWAY
        ]}
    except ValueError as e:
        print(f"[EROARE] Coloana lipsa in fisierul Excel: {e}")
        print(f"Coloane gasite in fisier: {header}")
        sys.exit(1)

    conn = get_connection()
    cur = conn.cursor()

    inserted, skipped, duplicates = 0, 0, 0
    seasons_seen = set()

    try:
        for row in ws.iter_rows(min_row=2, values_only=True):
            home_name = row[idx[COL_HOME]]
            away_name = row[idx[COL_AWAY]]
            raw_date = row[idx[COL_DATE]]
            season = row[idx[COL_SEASON]]

            if not home_name or not away_name or not raw_date:
                skipped += 1
                continue

            match_date = to_iso_date(raw_date)
            if match_date is None:
                skipped += 1
                continue

            home_team_id = get_or_create_team(cur, str(home_name).strip(), team_type, country)
            away_team_id = get_or_create_team(cur, str(away_name).strip(), team_type, country)

            home_goals = safe_int(row[idx[COL_HG]])
            away_goals = safe_int(row[idx[COL_AG]])
            odds_home = safe_float(row[idx[COL_ODDS_HOME]])
            odds_draw = safe_float(row[idx[COL_ODDS_DRAW]])
            odds_away = safe_float(row[idx[COL_ODDS_AWAY]])

            cur.execute("""
                INSERT OR IGNORE INTO matches
                    (competition_id, season, date, stage, home_team_id, away_team_id,
                     home_goals, away_goals, odds_home, odds_draw, odds_away)
                VALUES (?, ?, ?, NULL, ?, ?, ?, ?, ?, ?, ?)
            """, (competition_id, str(season) if season else None, match_date,
                  home_team_id, away_team_id, home_goals, away_goals,
                  odds_home, odds_draw, odds_away))

            if cur.rowcount == 0:
                duplicates += 1
            else:
                inserted += 1
                if season:
                    seasons_seen.add(str(season))

        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

    print(f"\nImport finalizat pentru '{competition_name}':")
    print(f"  - meciuri inserate: {inserted}")
    print(f"  - meciuri deja existente (sarite, fara duplicare): {duplicates}")
    print(f"  - randuri sarite (date/echipe lipsa): {skipped}")
    print(f"  - sezoane importate: {len(seasons_seen)} ({', '.join(sorted(seasons_seen))})")


if __name__ == "__main__":
    if len(sys.argv) != 5:
        print(__doc__)
        sys.exit(1)

    _, filepath, competition_name, competition_type, country = sys.argv
    country = None if country == "None" else country

    import_xlsx(filepath, competition_name, competition_type, country)
